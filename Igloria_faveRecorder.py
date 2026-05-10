import openpyxl as excel
import time
import os


os.system("cls")


workbook = excel.Workbook()
sheet = workbook.active


headers = ["ID", "First Name", "Last Name", "Birth Year", "Age"]
sheet.append(headers)


person_list = ["First", "Second", "Third"]


person_id = 0


def typing(text, speed):
    for letter in text:
        print(letter, end="", flush=True)
        time.sleep(speed)

for person in person_list:

    typing(f"{person} Favorite Person\n\n", 0.03)

    first_name = input("First Name: ")
    last_name = input("Last Name: ")
    birth_year = input("Birth Year: ")

    print()

    
    if not birth_year.isdigit():

        os.system("cls")

        typing("Birth year must be a number\n", 0.03)
        time.sleep(0.3)

        typing("Input not saved\n", 0.03)
        time.sleep(0.3)

        typing("Run the system again\n", 0.03)
        time.sleep(0.3)

        typing("Thank you!\n", 0.03)

        exit()

    
    person_id += 1

    
    age = 2026 - int(birth_year)

    
    sheet.append([
        f"0{person_id}",
        first_name,
        last_name,
        birth_year,
        age
    ])

    
    workbook.save("favorite_people.xlsx")

    os.system("cls")


typing("Saving file...", 0.2)

os.system("cls")

typing("...\n", 0.3)

os.system("cls")


typing("Favorite People List\n\n", 0.03)

saved_workbook = excel.load_workbook("favorite_people.xlsx")
saved_sheet = saved_workbook.active

for row in saved_sheet.iter_rows(values_only=True):
    print(row)


input("\nPress Enter to Exit... ")

os.system("cls")

typing("Thank you for using my system!", 0.03)
